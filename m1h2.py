# -*- coding: utf-8 -*-
"""
Created on Tue Nov  3 11:03:26 2020

@author: Kevin
"""

from openpyxl import Workbook
from openpyxl import load_workbook

#%%
wb = Workbook()
ws = wb.active

ws1 = wb.create_sheet("kerja")
ws2 = wb.create_sheet("kerja",0)
ws3 = wb.create_sheet("kerja",-1)

print(wb.sheetnames)

ws.title = "kerjabaru"
print(wb.sheetnames)

workaktif = wb['kerjabaru']
print(workaktif)

ws['B4'] = 4
ws.cell(row=4, column=2).value = 4

a = ['aku', 'suka', 'makan', 'bakso']

for row in  range(0,len(a)):
    ws.cell(row = row+1, column= 1).value = a[row]
    
for column in range(0,len(a)):
    ws.cell(row=1, column=column+1).value = a[column]
    
print(ws)

menu = [['hari'],['senin','nasi','ayam'],['selasa','susu'],['rabu','nasi goreng','ati ampela','jus apel'],['kamis', 'capcay', 'telur mata sapi']]

for item in menu:
    ws.append(item)

wb.save('coba.xlsx')
#%%
#baca excel 
data = load_workbook(filename="test.xlsx")
print(data.sheetnames)
sheet = data.active
print(sheet)
# ws.title = "nomor1"
# sheet = ws
print(sheet)
print(sheet['A1'])
print(sheet['A1'].value)
print(sheet.cell(row=1, column=1).value)

# sheet[“A3:D4”] akan mengambil data pada kota di dalam A3 sampai D4
# sheet[“A”] akan mengambil semua data pada kolom A
# sheet[1] akan mengambil semua data pada baris pertama
# sheet[“A:B”] akan mengambil semua data pada kolom A sampai kolom B
# sheet[1:4] akan mengambil semua data pada baris pertama sampai bari keempat.

sheet.iter_cols(min_row=1, max_row=4, min_col=1, max_col=4)
sheet.iter_rows(min_row=1, max_row=4, min_col=1, max_col=4)

#df to excel
# df.to_excel("namafile.xlsx")
#masukin df ke sheet tertentu
# df.to_excel("namafile.xlsx",sheet_name="nama_sheet")

#memasukan df ke excel yang sudah ada
"""
with pd.ExcelWriter('namafile.xlsx') as writer:
    hasil1.to_excel(writer, sheet_name='nama_sheet_1')
    hasil2.to_excel(writer, sheet_name='nama_sheet_2')
"""
#%%
from openpyxl import Workbook
from openpyxl.chart import BarChart, Series, Reference 
import csv

#inisiasi excel
wb = Workbook()
ws = wb.active 

#openFile
data = open("pulau_indonesia.csv")
rows = csv.reader(data, delimiter=',')

index = 0
for row in rows:
    data_clean = []
    for i  in row:
        try:
            i = int(i)
        except:
            pass
        data_clean.append(i)
    ws.append(data_clean)
    index +=1 
len_row = len(data_clean)

chart1 = BarChart()
chart1.type = "col"
chart1.style = 3
chart1.title = "Bar Chart"
chart1.y_axis.title = "Jumlah Pulau"
chart1.x_axis.title = "Nama Provinsi"

data = Reference(ws, min_col=3, min_row=1, max_row=index, max_col=len_row-1)
cats = Reference(ws, min_col=2, min_row=2, max_row = index)
chart1.height = 10
chart1.width = 30
chart1.add_data(data, titles_from_data=True)
chart1.set_categories(cats)
ws.add_chart(chart1,"G2")

wb.save("bar1.xlsx")